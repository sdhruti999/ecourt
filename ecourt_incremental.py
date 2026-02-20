
# import requests
# from bs4 import BeautifulSoup
# from openpyxl import Workbook, load_workbook
# import json
# from datetime import datetime
# import uuid
# import html as html_lib
# import re
# import time
# import os
# import random

# # ================= CONFIG =================

# CASE_LIST_URL = "https://services.ecourts.gov.in/ecourtindia_v6/?p=casestatus/submitFirNo"
# CASE_DETAIL_URL = "https://services.ecourts.gov.in/ecourtindia_v6/?p=home/viewHistory"

# EXCEL_FILE = "ecourts_incremental_output.xlsx"

# HEADERS = {
#     "accept": "application/json, text/javascript, */*; q=0.01",
#     "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
#     "origin": "https://services.ecourts.gov.in",
#     "referer": "https://services.ecourts.gov.in/",
#     "user-agent": "Mozilla/5.0",
#     "x-requested-with": "XMLHttpRequest",
# }

# FIR_SEARCH_PAYLOAD = {
#     "police_st_code": "120112",
#     "uniform_code": "11191044",
#     "state_code": "17",
#     "dist_code": "13",
#     "court_complex_code": "1170033",
#     "est_code": "null",
#     "case_status": "Both",
#     "ajax_req": "true",
#     "app_token": "229b64836be117cecc493c49b76cd3db2abc4d3e45736010bb6e2a41db15e976",
# }

# CASE_DETAIL_PAYLOAD = {
#     "court_code": "",
#     "state_code": "",
#     "dist_code": "",
#     "court_complex_code": "",
#     "case_no": "",
#     "cino": "",
#     "search_flag": "",
#     "search_by": "",
#     "ajax_req": "true",
#     "app_token": "",
# }

# # ================= EXCEL INIT =================

# def init_excel(headers):
#     if not os.path.exists(EXCEL_FILE):
#         wb = Workbook()
#         ws = wb.active
#         ws.append(headers)
#         wb.save(EXCEL_FILE)

# def append_excel(row, headers):
#     wb = load_workbook(EXCEL_FILE)
#     ws = wb.active
#     ws.append([row.get(h) for h in headers])
#     wb.save(EXCEL_FILE)

# # ================= FETCH CASE LIST =================

# def fetch_case_list():
#     res = requests.post(CASE_LIST_URL, headers=HEADERS, data=FIR_SEARCH_PAYLOAD)
#     res.raise_for_status()

#     soup = BeautifulSoup(res.json().get("case_data", ""), "html.parser")
#     cases = []

#     for a in soup.select("a[onclick^='viewHistory']"):
#         onclick = a.get("onclick")

#         m = re.search(
#             r"viewHistory\((\d+),'([^']+)',(\d+),'','([^']+)',(\d+),(\d+),(\d+),'([^']+)'\)",
#             onclick
#         )

#         if m:
#             cases.append({
#                 "case_no": m.group(1),
#                 "cino": m.group(2),
#                 "court_code": m.group(3),
#                 "search_flag": m.group(4),
#                 "state_code": m.group(5),
#                 "dist_code": m.group(6),
#                 "court_complex_code": m.group(7),
#                 "search_by": m.group(8),
#             })

#     return cases

# # ================= HELPERS =================

# def fetch_case_html(payload):
#     res = requests.post(CASE_DETAIL_URL, headers=HEADERS, data=payload)
#     res.raise_for_status()
#     return res.json()["data_list"]

# def extract_value(soup, label):
#     lbl = soup.find("label", string=lambda x: x and label in x)
#     if lbl:
#         td = lbl.find_parent("td")
#         if td:
#             nxt = td.find_next_sibling("td")
#             return nxt.get_text(strip=True) if nxt else None
#     return None

# def extract_case_type(soup):
#     for tr in soup.find_all("tr"):
#         tds = tr.find_all("td")
#         if len(tds) >= 2 and tds[0].get_text(strip=True).lower() == "case type":
#             return tds[1].get_text(strip=True)
#     return None

# def extract_fir_details(soup):
#     data = {}
#     table = soup.find("table", class_="FIR_details_table")
#     if not table:
#         return data

#     for tr in table.find_all("tr"):
#         tds = tr.find_all("td")
#         if len(tds) >= 2:
#             data[tds[0].get_text(strip=True)] = tds[1].get_text(strip=True)
#     return data

# # ================= PARSER =================

# def parse_case(html_text, payload):
#     soup = BeautifulSoup(html_lib.unescape(html_text), "html.parser")

#     row = {
#         "_id": uuid.uuid4().hex,
#         "Update Time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
#         "Court Name": soup.select_one("#chHeading").get_text(strip=True) if soup.select_one("#chHeading") else None,
#         "Case Type": extract_case_type(soup),
#         "Filing Date": extract_value(soup, "Filing Date"),
#         "Registration Date": extract_value(soup, "Registration Date"),
#         "First Hearing Date": extract_value(soup, "First Hearing Date"),
#         "Decision Date": extract_value(soup, "Decision Date"),
#         "Case Status": extract_value(soup, "Case Status"),
#         "Nature of Disposal": extract_value(soup, "Nature of Disposal"),
#     }

#     reg = extract_value(soup, "Registration Number")
#     if reg and "/" in reg:
#         row["Case Number"], row["Case Year"] = reg.split("/")
#     else:
#         row["Case Number"] = row["Case Year"] = None

#     fir = extract_fir_details(soup)
#     row["Police Station"] = fir.get("Police Station")
#     row["FIR Number"] = fir.get("FIR Number")
#     row["FIR Year"] = fir.get("Year")

#     row.update(payload)
#     row["status"] = "success"

#     return row

# # ================= MAIN =================

# def run():
#     cases = fetch_case_list()
#     headers_written = False
#     headers = None

#     print(f"Total cases found: {len(cases)}")

#     for idx, case in enumerate(cases, 1):
#         print(f"[{idx}/{len(cases)}] Fetching case_no={case['case_no']}")

#         payload = CASE_DETAIL_PAYLOAD.copy()
#         payload.update(case)
#         payload["app_token"] = FIR_SEARCH_PAYLOAD["app_token"]

#         html_data = fetch_case_html(payload)
#         row = parse_case(html_data, payload)

#         if not headers_written:
#             headers = list(row.keys())
#             init_excel(headers)
#             headers_written = True

#         append_excel(row, headers)
#         # time.sleep(1)
#         # time.sleep(random.uniform(2.5, 5.5))
#         sleep_time = random.choice([
#             random.uniform(2.5, 4.0),
#             random.uniform(4.0, 6.0),
#             random.uniform(6.0, 8.0),  # occasional long pause
#         ])
#         print(f"⏳ Sleeping {sleep_time:.2f} seconds")
#         time.sleep(sleep_time)


# # ================= ENTRY =================

# if __name__ == "__main__":
#     run()

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from datetime import datetime
import uuid
import html as html_lib
import re
import time
import os
import random

# ================= CONFIG =================

CASE_LIST_URL = "https://services.ecourts.gov.in/ecourtindia_v6/?p=casestatus/submitFirNo"
CASE_DETAIL_URL = "https://services.ecourts.gov.in/ecourtindia_v6/?p=home/viewHistory"

EXCEL_FILE = "ecourts_incremental_output.xlsx"

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Mozilla/5.0 (X11; Linux x86_64)",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)",
]

HEADERS = {
    "accept": "application/json, text/javascript, */*; q=0.01",
    "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
    "origin": "https://services.ecourts.gov.in",
    "referer": "https://services.ecourts.gov.in/",
    "x-requested-with": "XMLHttpRequest",
}

FIR_SEARCH_PAYLOAD = {
    "police_st_code": "120112",
    "uniform_code": "11191044",
    "state_code": "17",
    "dist_code": "13",
    "court_complex_code": "1170033",
    "est_code": "null",
    "case_status": "Both",
    "ajax_req": "true",
    "app_token": "229b64836be117cecc493c49b76cd3db2abc4d3e45736010bb6e2a41db15e976",
}

CASE_DETAIL_PAYLOAD = {
    "court_code": "",
    "state_code": "",
    "dist_code": "",
    "court_complex_code": "",
    "case_no": "",
    "cino": "",
    "search_flag": "",
    "search_by": "",
    "ajax_req": "true",
    "app_token": "",
}

# ================= SESSION =================

def create_session():
    session = requests.Session()
    HEADERS["user-agent"] = random.choice(USER_AGENTS)
    session.headers.update(HEADERS)
    return session

# ================= SAFE POST =================

def safe_post(session, url, payload, retries=3):
    for attempt in range(1, retries + 1):
        try:
            res = session.post(url, data=payload, timeout=30)

            if res.status_code in (403, 429):
                wait = 30 + attempt * 20
                print(f"⚠️ Block detected ({res.status_code}), sleeping {wait}s")
                time.sleep(wait)
                continue

            res.raise_for_status()
            return res

        except requests.RequestException as e:
            wait = 10 + attempt * 10
            print(f"⚠️ Network error: {e}, retrying in {wait}s")
            time.sleep(wait)

    return None

# ================= EXCEL =================

def init_excel(headers):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(EXCEL_FILE)

def append_excel(row, headers):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([row.get(h) for h in headers])
    wb.save(EXCEL_FILE)

# ================= HELPERS =================

def is_valid_html(html_text):
    return isinstance(html_text, str) and len(html_text.strip()) > 50

def extract_value(soup, label):
    lbl = soup.find("label", string=lambda x: x and label in x)
    if lbl:
        td = lbl.find_parent("td")
        if td:
            nxt = td.find_next_sibling("td")
            return nxt.get_text(strip=True) if nxt else None
    return None

def extract_case_type(soup):
    for tr in soup.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) >= 2 and tds[0].get_text(strip=True).lower() == "case type":
            return tds[1].get_text(strip=True)
    return None

def extract_fir_details(soup):
    data = {}
    table = soup.find("table", class_="FIR_details_table")
    if not table:
        return data

    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) >= 2:
            data[tds[0].get_text(strip=True)] = tds[1].get_text(strip=True)
    return data

# ================= PARSER =================

def parse_case(html_text, payload):
    if not is_valid_html(html_text):
        return {
            "_id": uuid.uuid4().hex,
            "Update Time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "status": "no_data",
            "error": "Empty or invalid HTML",
            **payload
        }

    soup = BeautifulSoup(html_lib.unescape(html_text), "html.parser")

    row = {
        "_id": uuid.uuid4().hex,
        "Update Time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Court Name": soup.select_one("#chHeading").get_text(strip=True) if soup.select_one("#chHeading") else None,
        "Case Type": extract_case_type(soup),
        "Filing Date": extract_value(soup, "Filing Date"),
        "Registration Date": extract_value(soup, "Registration Date"),
        "First Hearing Date": extract_value(soup, "First Hearing Date"),
        "Decision Date": extract_value(soup, "Decision Date"),
        "Case Status": extract_value(soup, "Case Status"),
        "Nature of Disposal": extract_value(soup, "Nature of Disposal"),
    }

    reg = extract_value(soup, "Registration Number")
    if reg and "/" in reg:
        row["Case Number"], row["Case Year"] = reg.split("/")
    else:
        row["Case Number"] = row["Case Year"] = None

    fir = extract_fir_details(soup)
    row["Police Station"] = fir.get("Police Station")
    row["FIR Number"] = fir.get("FIR Number")
    row["FIR Year"] = fir.get("Year")

    row.update(payload)
    row["status"] = "success"

    return row

# ================= FETCH CASE LIST =================

def fetch_case_list(session):
    res = safe_post(session, CASE_LIST_URL, FIR_SEARCH_PAYLOAD)
    if not res:
        return []

    soup = BeautifulSoup(res.json().get("case_data", ""), "html.parser")
    cases = []

    for a in soup.select("a[onclick^='viewHistory']"):
        onclick = a.get("onclick", "")

        m = re.search(
            r"viewHistory\((\d+),'([^']+)',(\d+),'','([^']+)',(\d+),(\d+),(\d+),'([^']+)'\)",
            onclick
        )

        if m:
            cases.append({
                "case_no": m.group(1),
                "cino": m.group(2),
                "court_code": m.group(3),
                "search_flag": m.group(4),
                "state_code": m.group(5),
                "dist_code": m.group(6),
                "court_complex_code": m.group(7),
                "search_by": m.group(8),
            })

    return cases

# ================= MAIN =================

def run():
    session = create_session()
    cases = fetch_case_list(session)

    print(f"Total cases found: {len(cases)}")

    headers_written = False
    headers = None

    for idx, case in enumerate(cases, 1):
        print(f"[{idx}/{len(cases)}] Fetching case_no={case['case_no']}")

        payload = CASE_DETAIL_PAYLOAD.copy()
        payload.update(case)
        payload["app_token"] = FIR_SEARCH_PAYLOAD["app_token"]

        res = safe_post(session, CASE_DETAIL_URL, payload)
        data = res.json() if res else {}
        html_data = data.get("data_list")

        row = parse_case(html_data, payload)

        if not headers_written:
            headers = list(row.keys())
            init_excel(headers)
            headers_written = True

        append_excel(row, headers)

        if row["status"] != "success":
            print(f"⚠️ No data for case_no={case['case_no']}")

        sleep_time = random.uniform(2.5, 5.5)
        time.sleep(sleep_time)

# ================= ENTRY =================

if __name__ == "__main__":
    run()
